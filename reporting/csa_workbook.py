"""CSA Workbook builder — template-based, data-only writer.

Layer 5 contract: **Excel = data, HTML = narrative.**

Copies the pre-built ``.xlsm`` template and writes **only** data values
into the existing sheets:

- ``0_Executive_Summary`` — engagement framing + assessment metrics + deterministic top risks
- ``1_30-60-90_Roadmap`` — phased remediation initiatives
- ``2_Control_Details`` — one row per assessed control (columns A–O)
  plus enrichment column P (Control Source)

Top risks are embedded in the Executive Summary sheet.  AI-generated
narrative (root cause, business impact, cascade effect) is rendered
in the HTML report only.

The template owns **all** visualisation: Dashboard formulas, charts,
conditional formatting, data validation, and VBA macros.  Python never
creates, modifies, or deletes sheets, formatting, formulas, or macros.

After saving, a ZIP-level restoration step re-injects any x14
extensions that openpyxl strips during its load / save cycle so the
workbook opens in Excel without corruption warnings.

Usage::

    from reporting.csa_workbook import build_csa_workbook
    build_csa_workbook(
        run_path="out/run.json",
        output_path="out/CSA_Workbook_v1.xlsm",
        why_payloads=[...],
    )
"""
from __future__ import annotations

import json
import os
import re
import shutil
import warnings
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════

def _load_json(path: str | None) -> dict:
    if not path or not Path(path).exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _safe_get(obj: Any, dotpath: str, default: Any = "") -> Any:
    for key in dotpath.split("."):
        if isinstance(obj, dict):
            obj = obj.get(key, {})
        else:
            return default
    return obj if obj != {} else default


def _join_list(value) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if v)
    return str(value) if value else ""


# ══════════════════════════════════════════════════════════════════
# Template + sheet constants
# ══════════════════════════════════════════════════════════════════

_TEMPLATE_DIR = Path(__file__).resolve().parent
_TEMPLATE_NAME = "Landing_Zone_Assessment.xlsm"
_TEMPLATE_PATH = _TEMPLATE_DIR / _TEMPLATE_NAME

# Sheet names — must match template exactly
_SHEET_EXEC     = "0_Executive_Summary"
_SHEET_ROADMAP  = "1_30-60-90_Roadmap"
_SHEET_CONTROLS = "2_Control_Details"

# Control Details layout (row 9 = headers, row 10+ = data)
_CD_HEADER_ROW = 9
_CD_DATA_START = 10

# Canonical status → workbook display.  Every ControlStatus has an entry.
_STATUS_MAP: dict[str, str] = {
    "Pass":             "Fulfilled",
    "Fail":             "Open",
    "Partial":          "Open",
    "Manual":           "Not verified",
    "NotApplicable":    "N/A",
    "NotVerified":      "Not verified",
    "SignalError":      "Not verified (Signal failure)",
    "EvaluationError":  "Not verified (Eval error)",
    # Legacy synonyms (workbook may receive pre-mapped values)
    "Fulfilled":    "Fulfilled",
    "Open":         "Open",
    "Not verified": "Not verified",
    "Not required": "Not required",
    "N/A":          "N/A",
}


def _map_status(raw: str) -> str:
    mapped = _STATUS_MAP.get(raw)
    if mapped is None:
        raise ValueError(f"Unmapped control status '{raw}' — add to _STATUS_MAP or fix evaluator")
    return mapped


# ══════════════════════════════════════════════════════════════════
# Data clearing
# ══════════════════════════════════════════════════════════════════

def _clear_data_rows(ws, start_row: int = _CD_DATA_START, max_col: int = 25):
    """Clear data rows without touching headers or table structure."""
    from openpyxl.cell.cell import MergedCell
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _write_control_detail_rows(
    ws,
    results: list[dict],
    checklist_lookup: dict[str, dict],
) -> int:
    """Populate 2_Control_Details starting at row 10.

    Columns A–O:
      A: ID  B: Design Area  C: Sub Area  D: WAF Pillar  E: Service
      F: Checklist item  G: Severity  H: Status  I: Comment
      J: AMMP  K: More info  L: Training
      M: Coverage: % Compliant  N: Coverage: Subs Affected  O: Coverage: Scope Level

    Coverage breakdown columns M–O show how widely a tenant-scoped finding
    manifests across input subscriptions.  They are NOT per-subscription
    maturity.  Enrichment appends P: Control Source after the data columns.
    One row per control — never per-subscription.

    Returns the number of rows written.
    """
    # ── Write / refresh column headers ────────────────────────────
    _HEADERS = [
        "ID", "Design Area", "Sub Area", "WAF Pillar", "Service",
        "Checklist item", "Severity", "Status", "Comment",
        "AMMP", "More info", "Training",
        "Coverage: % Compliant", "Coverage: Subs Affected", "Coverage: Scope Level",
    ]
    for col, hdr in enumerate(_HEADERS, start=1):
        ws.cell(row=_CD_HEADER_ROW, column=col, value=hdr)

    # Clear stale headers beyond the new layout
    for col in range(len(_HEADERS) + 1, 25):
        cell = ws.cell(row=_CD_HEADER_ROW, column=col)
        if cell.value:
            cell.value = None

    row = _CD_DATA_START

    for ctrl in results:
        cid = ctrl.get("control_id", "")
        cl = checklist_lookup.get(cid, {})

        ws.cell(row=row, column=1,  value=cl.get("id", ""))
        ws.cell(row=row, column=2,  value=cl.get(
            "category", ctrl.get("category", ctrl.get("section", ""))))
        ws.cell(row=row, column=3,  value=cl.get("subcategory", ""))
        ws.cell(row=row, column=4,  value=cl.get("waf", ""))
        ws.cell(row=row, column=5,  value=cl.get("service", ""))
        ws.cell(row=row, column=6,  value=cl.get(
            "text", ctrl.get("text", ctrl.get("question", ""))))
        ws.cell(row=row, column=7,  value=ctrl.get(
            "severity", cl.get("severity", "")))
        ws.cell(row=row, column=8,  value=_map_status(
            ctrl.get("status", "Manual")))

        # Comment / evidence
        evidence = ctrl.get("evidence", [])
        parts: list[str] = []
        notes = ctrl.get("notes", "")
        if notes:
            parts.append(notes)
        for ev in evidence[:2]:
            if isinstance(ev, dict):
                s = ev.get("summary", ev.get("resource_id", ""))
                if s:
                    parts.append(str(s)[:120])
        ws.cell(row=row, column=9, value="\n".join(parts))

        ws.cell(row=row, column=10, value="")                       # AMMP
        ws.cell(row=row, column=11, value=cl.get("link", ""))       # Learn link
        ws.cell(row=row, column=12, value=cl.get("training", ""))   # Training

        # ── Coverage breakdown columns (M–P) ──────────────────────
        # These are tenant-scoped coverage metrics, NOT per-subscription
        # maturity.  They show how many input subscriptions exhibit a finding.
        # % Compliant: e.g. "85.0%" or "17/100 compliant"
        cov_display = ctrl.get("coverage_display")
        cov_pct = ctrl.get("coverage_pct")
        if cov_display:
            ws.cell(row=row, column=13, value=cov_display)
        elif cov_pct is not None:
            ws.cell(row=row, column=13, value=f"{cov_pct}%")
        else:
            ws.cell(row=row, column=13, value="")

        # Coverage: Subs Affected — how many input subs exhibit this finding (not per-sub maturity)
        subs_affected = ctrl.get("subscriptions_affected")
        subs_assessed = ctrl.get("subscriptions_assessed")
        if subs_affected is not None and subs_assessed:
            ws.cell(row=row, column=14,
                    value=f"{subs_affected}/{subs_assessed}")
        else:
            ws.cell(row=row, column=14, value="")

        # Coverage: Scope Level — where the finding manifests, not the assessment scope
        ws.cell(row=row, column=15, value=ctrl.get("scope_level", ""))

        row += 1

    return row - _CD_DATA_START


# ══════════════════════════════════════════════════════════════════
# 0_Executive_Summary  — populate existing rows
# ══════════════════════════════════════════════════════════════════

def _populate_executive_summary(ws, run: dict) -> None:
    """Write values into the existing Executive Summary layout.

    Layer 5 contract: Excel = data, HTML = narrative.
    Top risks (rows 17+) are populated from deterministic risk scoring,
    NOT from AI-generated business risk language.

    Template layout (column A = labels, column B = values):
      Row 1:  CSA ENGAGEMENT FRAMING (title — leave as-is)
      Row 2:  Engagement Objective | <text>
      Row 3:  Key Message | <text>
      Row 4:  Customer Outcome | <text>
      Row 6:  ASSESSMENT METRICS (title — leave as-is)
      Row 7:  Tenant ID | <value>
      Row 8:  Assessment Date | <value>
      Row 9:  Enterprise-Scale Ready | <value>
      Row 10: Tenant Maturity | <value>
      Row 11: Data-Driven Controls | <count>
      Row 12: Requires Customer Input | <count>
      Row 13: Subscriptions (Input Scope) | <count>
      Row 15: Top Risks — Deterministic (title — leave as-is)
      Row 16: Control | Risk Tier | Score (sub-header — leave as-is)
      Row 17+: deterministic risk data rows
    """
    from engine.risk_scoring import build_risk_overview

    es = run.get("executive_summary", {})
    scoring = run.get("scoring", {})
    ec = run.get("execution_context", {})
    ai = run.get("ai", {})
    esr = ai.get("enterprise_scale_readiness", {})
    results = run.get("results", [])
    total_controls = len(results)

    # ── Engagement framing (factual, not AI-generated) ────────────
    objective = (
        "Assess the customer's Azure landing zone maturity, identify "
        "critical gaps, and deliver a prioritised 30-60-90 remediation "
        "roadmap aligned to Microsoft Cloud Adoption Framework."
    )

    # Derive top risks deterministically for the key message
    risk_overview = build_risk_overview(results)
    crit_count = risk_overview["summary"]["critical_count"]
    high_count = risk_overview["summary"]["high_count"]
    total_risk = risk_overview["summary"]["total_risk_count"]

    key_message = (
        f"This assessment evaluated {total_controls} controls across the "
        f"tenant using live platform telemetry. Deterministic risk scoring "
        f"identified {total_risk} at-risk controls "
        f"({crit_count} critical, {high_count} high). "
        f"The roadmap ties each action to specific controls, making "
        f"every recommendation defensible and auditable."
    )
    customer_outcome = (
        "A data-driven workbook the customer owns — with scored controls, "
        "a traceable remediation plan, and Microsoft Learn references — "
        "enabling them to drive implementation with or without further "
        "Microsoft engagement."
    )

    ws.cell(row=2, column=2, value=objective)
    ws.cell(row=3, column=2, value=key_message)
    ws.cell(row=4, column=2, value=customer_outcome)

    # ── Assessment metrics ────────────────────────────────────────
    ws.cell(row=7, column=2, value=ec.get("tenant_id", "Unknown"))
    ws.cell(row=8, column=2, value=run.get("meta", {}).get("timestamp", ""))

    ready = esr.get("ready_for_enterprise_scale", False)
    score = esr.get("readiness_score", "")
    ws.cell(row=9, column=2, value="Yes" if ready else f"No  (score: {score})")

    # Scope model: maturity is tenant-wide, never per-subscription
    ws.cell(row=10, column=1, value="Tenant Maturity")
    maturity = scoring.get('overall_maturity_percent')
    ws.cell(row=10, column=2,
            value=f"{maturity}%" if maturity is not None else "Unavailable")

    data_driven = sum(
        1 for r in results
        if r.get("status") in ("Pass", "Fail", "Partial", "Fulfilled", "Open")
        and r.get("signal_used")
    )
    ws.cell(row=11, column=2, value=data_driven)
    ws.cell(row=12, column=2, value=total_controls - data_driven)

    # Subscriptions are inputs (data sources), not evaluation units
    ws.cell(row=13, column=1, value="Subscriptions (Input Scope)")
    ws.cell(row=13, column=2, value=ec.get("subscription_count_visible", ""))

    # ── Top risks table (row 15+) — deterministic, data-only ─────
    # Sub-headers
    ws.cell(row=15, column=1, value="Top Risks — Deterministic")
    ws.cell(row=16, column=1, value="Control")
    ws.cell(row=16, column=2, value="Risk Tier")
    ws.cell(row=16, column=3, value="Score")

    # Flatten top-N from tiers (Critical first, then High)
    top_controls: list[dict] = []
    for tier_name in ("Critical", "High", "Medium"):
        top_controls.extend(risk_overview["tiers"].get(tier_name, []))
    top_controls = top_controls[:10]  # cap at 10 rows

    row = 17
    for ctrl in top_controls:
        ws.cell(row=row, column=1, value=ctrl.get("text", ""))
        ws.cell(row=row, column=2, value=ctrl.get("risk_tier", ""))
        ws.cell(row=row, column=3, value=ctrl.get("risk_score", ""))
        row += 1


# ══════════════════════════════════════════════════════════════════
# 1_30-60-90_Roadmap  — populate existing rows
# ══════════════════════════════════════════════════════════════════

def _populate_roadmap(ws, run: dict) -> int:
    """Write values into the existing Roadmap layout (row 1 = headers).

    Columns: Phase | Action | Initiative ID | CAF Discipline | Owner |
             Success Criteria | Dependencies | Related Controls | Related Risks

    Returns the number of rows written.
    """
    tr = run.get("transformation_roadmap", {})
    roadmap = tr.get("roadmap_30_60_90", {})
    tp = run.get("transformation_plan", {})
    init_lookup: dict[str, dict] = {
        i.get("initiative_id", ""): i
        for i in tp.get("initiatives", [])
        if i.get("initiative_id")
    }

    phase_map = {"30_days": "30 Days", "60_days": "60 Days", "90_days": "90 Days"}
    row = 2
    for phase_key, phase_label in phase_map.items():
        for item in roadmap.get(phase_key, []):
            iid = item.get("initiative_id", "")
            init_detail = init_lookup.get(iid, {})
            ws.cell(row=row, column=1, value=phase_label)
            ws.cell(row=row, column=2, value=item.get("action", ""))
            ws.cell(row=row, column=3, value=iid)
            ws.cell(row=row, column=4, value=item.get("caf_discipline", ""))
            ws.cell(row=row, column=5, value=item.get("owner_role", ""))
            ws.cell(row=row, column=6, value=item.get("success_criteria", ""))
            ws.cell(row=row, column=7, value=_join_list(
                item.get("dependency_on", [])))
            ws.cell(row=row, column=8, value=_join_list(
                init_detail.get("controls", [])))
            ws.cell(row=row, column=9, value="")  # filled by cross-ref below
            row += 1

    _cross_ref_roadmap_risks(ws, run, start_row=2, end_row=row - 1)
    return row - 2


def _cross_ref_roadmap_risks(
    ws, run: dict, start_row: int, end_row: int,
) -> None:
    """Fill column I (Related Risks) by matching initiative controls to risk tiers.

    Layer 5: uses deterministic risk scoring, not AI-generated risk titles.
    For each roadmap row, finds related controls from column H and reports
    their highest deterministic risk tier.
    """
    from engine.risk_scoring import build_risk_overview

    results = run.get("results", [])
    overview = build_risk_overview(results)

    # Build control_id → risk_tier lookup
    risk_lookup: dict[str, str] = {}
    for tier_name, controls in overview["tiers"].items():
        for ctrl in controls:
            sid = ctrl.get("short_id", "")
            if sid:
                risk_lookup[sid] = tier_name

    for r in range(start_row, end_row + 1):
        related_ctrls = str(ws.cell(row=r, column=8).value or "")
        if not related_ctrls:
            continue
        ctrl_ids = {c.strip()[:8] for c in related_ctrls.replace(";", ",").split(",")}
        matched_tiers: set[str] = set()
        for cid in ctrl_ids:
            tier = risk_lookup.get(cid)
            if tier:
                matched_tiers.add(tier)
        if matched_tiers:
            # Show highest tier first
            tier_order = ["Critical", "High", "Medium", "Hygiene"]
            sorted_tiers = sorted(matched_tiers, key=lambda t: tier_order.index(t) if t in tier_order else 99)
            ws.cell(row=r, column=9, value="; ".join(sorted_tiers))



# ══════════════════════════════════════════════════════════════════
# Signal integrity validation
# ══════════════════════════════════════════════════════════════════

class SignalIntegrityError(RuntimeError):
    """Raised when no platform signals were collected — report would be hollow."""


def validate_signal_integrity(run: dict, *, allow_demo: bool = False) -> dict:
    """Verify the run contains live platform signals before rendering.

    Returns a ``provenance`` dict with scan duration, API counts, and
    signal inventory.  Raises ``SignalIntegrityError`` if signal counts
    are zero and ``allow_demo`` is False.
    """
    telemetry = run.get("telemetry", {})
    is_live = telemetry.get("live_run", False)
    sig_avail = run.get("signal_availability", {})
    results = run.get("results", [])

    # Count signals that actually returned data — use None when absent
    rg_queries = telemetry.get("rg_query_count")
    arm_calls = telemetry.get("arm_call_count")
    signals_fetched = telemetry.get("signals_fetched")
    total_api_calls = (rg_queries or 0) + (arm_calls or 0)

    # Signal inventory from availability matrix
    signal_inventory: dict[str, int] = {}
    for category, sigs in sig_avail.items():
        if isinstance(sigs, list):
            signal_inventory[category] = len(sigs)

    # Data-driven controls (have a signal_used value)
    data_driven = sum(1 for r in results if r.get("signal_used"))

    provenance = {
        "live": is_live,
        "statement": (
            "This report was generated from live platform telemetry. "
            "No questionnaire or Excel input was used."
        ) if is_live else (
            "Demo Mode \u2014 No live telemetry. "
            "Metrics shown are from cached or sample data."
        ),
        "scan_duration_sec": telemetry.get("assessment_duration_sec"),
        "api_calls_total": total_api_calls if is_live else None,
        "rg_queries": rg_queries,
        "arm_calls": arm_calls,
        "signals_fetched": signals_fetched,
        "signals_cached": telemetry.get("signals_cached"),
        "signal_errors": telemetry.get("signal_errors"),
        "signal_inventory": signal_inventory,
        "signal_categories": len(signal_inventory),
        "data_driven_controls": data_driven,
        "total_controls": len(results),
    }

    # Gate: abort if no live signals and not demo
    if total_api_calls == 0 and data_driven == 0 and not allow_demo:
        raise SignalIntegrityError(
            "ABORT: Platform signal counts are zero. "
            "No live telemetry was collected — cannot generate a credible report. "
            f"(rg_queries={rg_queries}, arm_calls={arm_calls}, "
            f"data_driven_controls={data_driven})"
        )

    return provenance


# ══════════════════════════════════════════════════════════════════
# ZIP-level extension restoration
# ══════════════════════════════════════════════════════════════════

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _zip_sheet_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """Return ``{sheet_name: zip_path}`` from workbook.xml + rels."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_target: dict[str, str] = {}
    for r in rels:
        t = r.get("Target", "")
        # Normalise to ZIP-entry path (no leading /, relative to xl/)
        t = t.lstrip("/")
        if not t.startswith("xl/"):
            t = "xl/" + t
        rid = r.get("Id", "")
        if rid:
            rid_target[rid] = t

    result: dict[str, str] = {}
    for sh in wb.findall(f".//{{{_NS_MAIN}}}sheet"):
        name = sh.get("name", "")
        rid = sh.get(f"{{{_NS_REL}}}id")
        if name and rid and rid in rid_target:
            result[name] = rid_target[rid]
    return result


def _extract_ws_extlst(data: bytes) -> bytes | None:
    """Extract the worksheet-level ``<extLst>…</extLst>`` block.

    The worksheet-level extLst is always the last child element before
    ``</worksheet>``.  Handles nested extLst elements correctly.
    """
    ws_end = data.rfind(b"</worksheet>")
    if ws_end == -1:
        return None
    region = data[:ws_end]

    start = region.rfind(b"<extLst")
    if start == -1:
        return None

    # Walk forward to find the matching </extLst> (handles nesting)
    depth = 0
    pos = start
    OPEN = b"<extLst"
    CLOSE = b"</extLst>"
    # Move past the initial tag
    scan_from = start + len(OPEN)
    while scan_from < ws_end:
        next_open = data.find(OPEN, scan_from)
        next_close = data.find(CLOSE, scan_from)

        if next_close == -1:
            return None

        if next_open != -1 and next_open < next_close:
            depth += 1
            scan_from = next_open + len(OPEN)
        else:
            if depth == 0:
                return data[start : next_close + len(CLOSE)]
            depth -= 1
            scan_from = next_close + len(CLOSE)
    return None


def _extract_ns_decls(data: bytes) -> list[bytes]:
    """Extract ``xmlns:*`` declarations from the root element tag."""
    # Skip XML declaration if present
    if data.startswith(b"<?"):
        decl_end = data.find(b"?>")
        root_start = data.find(b"<", decl_end + 2)
    else:
        root_start = 0
    root_end = data.find(b">", root_start)
    if root_end == -1:
        return []
    root_tag = data[root_start:root_end]
    return re.findall(rb'xmlns:\w+="[^"]*"', root_tag)


def _extract_mc_ignorable(data: bytes) -> bytes | None:
    """Extract ``mc:Ignorable`` attribute value from the root tag."""
    if data.startswith(b"<?"):
        decl_end = data.find(b"?>")
        root_start = data.find(b"<", decl_end + 2)
    else:
        root_start = 0
    root_end = data.find(b">", root_start)
    if root_end == -1:
        return None
    root_tag = data[root_start:root_end]
    m = re.search(rb'mc:Ignorable="([^"]*)"', root_tag)
    return m.group(1) if m else None


def _root_tag_end(data: bytes) -> int:
    """Return the byte offset of the first ``>`` in the root element."""
    if data.startswith(b"<?"):
        decl_end = data.find(b"?>")
        root_start = data.find(b"<", decl_end + 2)
    else:
        root_start = 0
    return data.find(b">", root_start)


def _restore_extensions(template_path: str, output_path: str) -> int:
    """Re-inject x14 extensions that openpyxl strips on load / save.

    For every worksheet present in **both** the template and the output,
    copies the ``<extLst>`` block (x14 conditional formatting, data
    validation, etc.) from the template back into the output.  Also
    ensures the required ``xmlns:*`` and ``mc:Ignorable`` declarations
    exist on the ``<worksheet>`` root element.

    The output ZIP is rewritten in-place (via a temp file).
    Returns the number of sheets patched.
    """
    # ── Read template sheet data ──────────────────────────────────
    with zipfile.ZipFile(template_path, "r") as zt:
        tpl_map = _zip_sheet_map(zt)
        tpl_data: dict[str, bytes] = {}
        for name, path in tpl_map.items():
            try:
                tpl_data[name] = zt.read(path)
            except KeyError:
                pass

    # ── Read all output ZIP entries (preserving order) ────────────
    with zipfile.ZipFile(output_path, "r") as zo:
        out_map = _zip_sheet_map(zo)
        entry_order = zo.namelist()
        entries: dict[str, bytes] = {n: zo.read(n) for n in entry_order}

    # ── Patch each sheet ──────────────────────────────────────────
    patched = 0
    for sheet_name, tpl_bytes in tpl_data.items():
        if sheet_name not in out_map:
            continue
        out_path = out_map[sheet_name]
        if out_path not in entries:
            continue

        extlst = _extract_ws_extlst(tpl_bytes)
        if extlst is None:
            continue

        out_bytes = entries[out_path]

        # Remove any partial extLst openpyxl may have left
        existing = _extract_ws_extlst(out_bytes)
        if existing:
            out_bytes = out_bytes.replace(existing, b"")

        # Inject template extLst before </worksheet>
        ws_end = out_bytes.rfind(b"</worksheet>")
        out_bytes = out_bytes[:ws_end] + extlst + b"\n" + out_bytes[ws_end:]

        # Ensure namespace declarations from template are present
        tpl_ns = _extract_ns_decls(tpl_bytes)
        for ns_decl in tpl_ns:
            rte = _root_tag_end(out_bytes)
            if ns_decl not in out_bytes[:rte]:
                out_bytes = out_bytes[:rte] + b" " + ns_decl + out_bytes[rte:]

        # Ensure mc:Ignorable matches the template
        tpl_mc = _extract_mc_ignorable(tpl_bytes)
        if tpl_mc:
            out_mc = _extract_mc_ignorable(out_bytes)
            if out_mc != tpl_mc:
                rte = _root_tag_end(out_bytes)
                root_tag = out_bytes[:rte]
                if out_mc:
                    root_tag = root_tag.replace(
                        b'mc:Ignorable="' + out_mc + b'"',
                        b'mc:Ignorable="' + tpl_mc + b'"',
                    )
                else:
                    root_tag += b' mc:Ignorable="' + tpl_mc + b'"'
                out_bytes = root_tag + out_bytes[rte:]

        entries[out_path] = out_bytes
        patched += 1

    # ── Rewrite the ZIP ───────────────────────────────────────────
    if patched > 0:
        tmp = output_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in entry_order:
                if name in entries:
                    zout.writestr(name, entries[name])
        os.replace(tmp, output_path)

    return patched


# ══════════════════════════════════════════════════════════════════
# Main builder
# ══════════════════════════════════════════════════════════════════

def build_csa_workbook(
    run_path: str = "out/run.json",
    target_path: str = "out/target_architecture.json",
    output_path: str = "out/CSA_Workbook_v1.xlsm",
    why_payloads: list[dict] | None = None,
    template_path: str | None = None,
) -> str:
    """Build the CSA workbook: copy template, populate all 4 data sheets.

    This function is a **data injector** — it writes only values into the
    template's existing data ranges.  It performs NO scoring, NO inference,
    NO sheet creation/deletion, and NO formatting changes.

    Layer 5 contract: Excel = data, HTML = narrative.
    Risk content is derived from ``build_risk_overview()`` — no AI narrative.

    Sheets populated:
      ``0_Executive_Summary`` — engagement framing + metrics + deterministic top risks
      ``1_30-60-90_Roadmap`` — phased initiatives
      ``2_Control_Details`` — one row per control (A–U) + enrichment (V–Y)
    """
    run = _load_json(run_path)

    # ── Resolve template ──────────────────────────────────────────
    tpl = Path(template_path) if template_path else _TEMPLATE_PATH
    if not tpl.exists():
        raise FileNotFoundError(
            f"Template not found: {tpl}\n"
            f"Place {_TEMPLATE_NAME} in {_TEMPLATE_DIR}/"
        )

    # ── Log provenance before rendering ───────────────────────────
    ec = run.get("execution_context", {})
    telem = run.get("telemetry", {})
    is_live = telem.get("live_run", False)
    print("  ┌─ Workbook Provenance ────────────────┐")
    print(f"  │ mode:                   {'Live' if is_live else 'Demo / Cached'}")
    print(f"  │ tenant_id:              {ec.get('tenant_id', 'N/A')}")
    print(f"  │ subscription_count:     {ec.get('subscription_count_visible', 'N/A')}")
    if is_live:
        print(f"  │ rg_queries:             {telem.get('rg_query_count', 'N/A')}")
        print(f"  │ arm_calls:              {telem.get('arm_call_count', 'N/A')}")
        print(f"  │ signals_fetched:        {telem.get('signals_fetched', 'N/A')}")
        print(f"  │ scan_duration:          {telem.get('assessment_duration_sec', 'N/A')}s")
    else:
        print("  │ telemetry:              Not available (demo/cached run)")
    print("  └─────────────────────────────────────────┘")

    # ── Copy template → output (byte-for-byte) ───────────────────
    out = Path(output_path)
    if out.suffix.lower() != ".xlsm":
        out = out.with_suffix(".xlsm")
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(tpl), str(out))

    # ── Single openpyxl pass (suppress extension warnings) ────────
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(str(out), keep_vba=True)

    # NEVER delete sheets — template owns all structure
    # Verify expected sheets exist
    for sheet_name in [_SHEET_EXEC, _SHEET_ROADMAP, _SHEET_CONTROLS]:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Sheet '{sheet_name}' not found in template — skipping")

    # ── Load ALZ checklist for rich per-control fields ────────────
    checklist_lookup: dict[str, dict] = {}
    try:
        from alz.loader import load_alz_checklist
        cl = load_alz_checklist()
        for item in cl.get("items", []):
            guid = item.get("guid", "")
            if guid:
                checklist_lookup[guid] = item
    except Exception:
        pass

    results = run.get("results", [])

    # ── 0_Executive_Summary ───────────────────────────────────────
    if _SHEET_EXEC in wb.sheetnames:
        _populate_executive_summary(wb[_SHEET_EXEC], run)
        print("  ✓ 0_Executive_Summary populated")

    # ── 1_30-60-90_Roadmap ────────────────────────────────────────
    if _SHEET_ROADMAP in wb.sheetnames:
        ws_rm = wb[_SHEET_ROADMAP]
        _clear_data_rows(ws_rm, start_row=2, max_col=9)
        n_roadmap = _populate_roadmap(ws_rm, run)
        print(f"  ✓ 1_30-60-90_Roadmap: {n_roadmap} initiatives")

    # ── 2_Control_Details (primary data sheet) ────────────────────
    if _SHEET_CONTROLS in wb.sheetnames:
        ws_cd = wb[_SHEET_CONTROLS]
        _clear_data_rows(ws_cd, start_row=_CD_DATA_START)
        n_controls = _write_control_detail_rows(ws_cd, results, checklist_lookup)

        # Enrichment in the same open workbook (no second load/save)
        try:
            from reporting.enrich import enrich_open_worksheet
            e_stats = enrich_open_worksheet(ws_cd)
            print(
                f"  ✓ 2_Control_Details: {n_controls} controls "
                f"({e_stats.get('alz', 0)} ALZ, "
                f"{e_stats.get('derived', 0)} derived)"
            )
        except Exception as e:
            print(f"  ✓ 2_Control_Details: {n_controls} controls "
                  f"(enrichment skipped: {e})")
    else:
        print(f"  ⚠ Sheet '{_SHEET_CONTROLS}' not found — skipping")

    # ── Save ──────────────────────────────────────────────────────
    try:
        wb.save(str(out))
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        fallback = out.with_name(f"{out.stem}_{ts}.xlsm")
        wb.save(str(fallback))
        print(f"  ⚠ Saved as {fallback.name} (original locked)")
        out = fallback

    # ── Restore x14 extensions stripped by openpyxl ───────────────
    patched = _restore_extensions(str(tpl), str(out))
    if patched:
        print(f"  ✓ Extensions restored for {patched} sheet(s)")

    print(f"  ✓ CSA workbook → {out}")
    return str(out)
